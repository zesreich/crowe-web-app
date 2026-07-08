import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import os

excel_path = r"C:\Users\Mert\Desktop\set\B10_OD_Dosya_Incelemesi.xlsx"

print(f"📂 Excel dosyası okunuyor: {excel_path}")

# Load workbook with data_only=False to get formulas and validations
try:
    wb = load_workbook(excel_path, data_only=False)
    print(f"✅ Workbook yüklendi. Sayfa sayısı: {len(wb.sheetnames)}")
except Exception as e:
    print(f"❌ Hata: {e}")
    exit(1)

# Dictionary to store all data
analysis_data = {
    'sheets': [],
    'data_validations': []
}

# Analyze each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Read sheet data with pandas
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
        sheet_data = {
            'name': sheet_name,
            'rows': len(df),
            'columns': len(df.columns) if len(df) > 0 else 0,
            'data': df.fillna('').astype(str).values.tolist()
        }
    except Exception as e:
        sheet_data = {
            'name': sheet_name,
            'rows': ws.max_row,
            'columns': ws.max_column,
            'data': [],
            'error': str(e)
        }
    
    # Find data validations (dropdown/çoktan seçmeli hücreler)
    validations = []
    validation_cells = {}  # Track which cells have validations
    
    try:
        for dv in ws.data_validations.dataValidation:
            # Get cells that have this validation
            for cell_range in dv.ranges:
                # Convert range to list of cells
                cells_in_range = []
                if hasattr(cell_range, 'cells'):
                    cells_in_range = list(cell_range.cells)
                else:
                    # Parse range string like "A1:B10"
                    range_str = str(cell_range)
                    if ':' in range_str:
                        start_cell, end_cell = range_str.split(':')
                        start_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, start_cell)))
                        start_row = int(''.join(filter(str.isdigit, start_cell)))
                        end_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, end_cell)))
                        end_row = int(''.join(filter(str.isdigit, end_cell)))
                        
                        for row in range(start_row, end_row + 1):
                            for col in range(start_col, end_col + 1):
                                cell_ref = f"{get_column_letter(col)}{row}"
                                cells_in_range.append(cell_ref)
                    else:
                        cells_in_range = [range_str]
                
                for cell_ref in cells_in_range:
                    try:
                        cell = ws[cell_ref]
                        cell_value = cell.value if cell.value is not None else ''
                        
                        # Get validation type and values
                        validation_info = {
                            'cell': cell_ref,
                            'row': cell.row,
                            'col': cell.column,
                            'value': str(cell_value),
                            'type': dv.type if hasattr(dv, 'type') else 'unknown',
                            'formula1': dv.formula1 if hasattr(dv, 'formula1') else None,
                            'formula2': dv.formula2 if hasattr(dv, 'formula2') else None,
                        }
                        
                        # Try to get list values if it's a list type
                        if dv.type == 'list' and dv.formula1:
                            list_values = []
                            formula = dv.formula1.replace('$', '').replace('"', '')
                            
                            # Check if it's a range reference
                            if '!' in formula or ':' in formula:
                                try:
                                    if '!' in formula:
                                        ref_sheet_name, ref_range = formula.split('!')
                                        ref_ws = wb[ref_sheet_name]
                                    else:
                                        ref_ws = ws
                                        ref_range = formula
                                    
                                    # Parse range
                                    if ':' in ref_range:
                                        start, end = ref_range.split(':')
                                        start_col_idx = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, start)))
                                        start_row_idx = int(''.join(filter(str.isdigit, start)))
                                        end_col_idx = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, end)))
                                        end_row_idx = int(''.join(filter(str.isdigit, end)))
                                        
                                        for r in range(start_row_idx, end_row_idx + 1):
                                            for c in range(start_col_idx, end_col_idx + 1):
                                                val = ref_ws.cell(row=r, column=c).value
                                                if val is not None and str(val).strip():
                                                    list_values.append(str(val).strip())
                                    else:
                                        val = ref_ws[ref_range].value
                                        if val is not None:
                                            list_values.append(str(val))
                                except Exception as e:
                                    list_values = [formula]
                            else:
                                # Direct list values
                                if ';' in formula:
                                    list_values = [v.strip() for v in formula.split(';') if v.strip()]
                                elif ',' in formula:
                                    list_values = [v.strip() for v in formula.split(',') if v.strip()]
                                else:
                                    list_values = [formula] if formula else []
                            
                            validation_info['list_values'] = list_values
                        
                        if cell_ref not in validation_cells:
                            validation_cells[cell_ref] = True
                            validations.append(validation_info)
                    except Exception as e:
                        print(f"⚠️ Hücre işlenirken hata: {cell_ref} - {e}")
                        continue
    except Exception as e:
        print(f"⚠️ Validation okuma hatası ({sheet_name}): {e}")
    
    sheet_data['validations'] = validations
    analysis_data['sheets'].append(sheet_data)

# Save analysis to JSON
output_json = json.dumps(analysis_data, indent=2, ensure_ascii=False, default=str)
with open('excel_analysis.json', 'w', encoding='utf-8') as f:
    f.write(output_json)

print("✅ Excel analizi tamamlandı. excel_analysis.json dosyası oluşturuldu.")
print(f"📊 Toplam {len(analysis_data['sheets'])} sayfa analiz edildi.")

