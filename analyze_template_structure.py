#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import json

excel_path = r"C:\Users\Mert\Desktop\set\B10_OD_Dosya_Incelemesi.xlsx"

print("📂 Template Excel yapısı analiz ediliyor...")

try:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # Analyze B10_FileReview sheet
    if 'B10_FileReview' in wb.sheetnames:
        ws = wb['B10_FileReview']
        print(f"\n📋 B10_FileReview Sayfası:")
        print(f"   Satır sayısı: {ws.max_row}")
        print(f"   Sütun sayısı: {ws.max_column}")
        
        # Find header row
        header_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            if any(val and 'Referans' in str(val) for val in row_values):
                header_row = row_idx
                print(f"\n   Header satırı: {row_idx}")
                print(f"   Header değerleri: {row_values[:10]}")
                break
        
        # Find data start row
        data_start_row = header_row + 1 if header_row else 2
        print(f"\n   Veri başlangıç satırı: {data_start_row}")
        
        # Analyze first few data rows
        print(f"\n   İlk 5 veri satırı:")
        for row_idx in range(data_start_row, min(data_start_row + 5, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx][:7]]
            print(f"   Satır {row_idx}: {row_values}")
        
        # Find column positions
        if header_row:
            header_values = [cell.value for cell in ws[header_row][:10]]
            column_map = {}
            for idx, val in enumerate(header_values):
                if val:
                    val_str = str(val).lower()
                    if 'referans' in val_str and 'kod' in val_str:
                        column_map['ref_code'] = idx + 1
                    elif 'soru' in val_str or 'no' in val_str:
                        column_map['question_num'] = idx + 1
                    elif 'prosedür' in val_str or 'procedure' in val_str:
                        column_map['procedure'] = idx + 1
                    elif 'açıklama' in val_str or 'explanation' in val_str:
                        column_map['explanation'] = idx + 1
                    elif 'sonuç' in val_str or 'result' in val_str:
                        column_map['result'] = idx + 1
                    elif 'risk' in val_str and 'atama' in val_str:
                        column_map['risk_assignment'] = idx + 1
                    elif 'referans' in val_str and 'kod' not in val_str:
                        column_map['reference'] = idx + 1
            
            print(f"\n   Sütun haritası: {column_map}")
            
            # Save to JSON
            structure = {
                'header_row': header_row,
                'data_start_row': data_start_row,
                'column_map': column_map,
                'max_row': ws.max_row,
                'max_column': ws.max_column
            }
            
            with open('html/template_column_map.json', 'w', encoding='utf-8') as f:
                json.dump(structure, f, ensure_ascii=False, indent=2)
            
            print(f"\n✅ Yapı bilgisi kaydedildi: html/template_column_map.json")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    import traceback
    traceback.print_exc()


