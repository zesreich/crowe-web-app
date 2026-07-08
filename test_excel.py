import sys
import traceback

try:
    print("Test başlıyor...")
    import openpyxl
    print("openpyxl import edildi")
    
    excel_path = r"C:\Users\Mert\Desktop\set\B10_OD_Dosya_Incelemesi.xlsx"
    print(f"Dosya yolu: {excel_path}")
    
    import os
    if os.path.exists(excel_path):
        print("✅ Dosya bulundu")
    else:
        print("❌ Dosya bulunamadı")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    print(f"✅ Workbook yüklendi")
    print(f"Sayfa sayısı: {len(wb.sheetnames)}")
    print(f"Sayfalar: {wb.sheetnames}")
    
    # Test first sheet
    if len(wb.sheetnames) > 0:
        ws = wb[wb.sheetnames[0]]
        print(f"İlk sayfa: {ws.title}, Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # Check validations
        validation_count = len(ws.data_validations.dataValidation) if hasattr(ws.data_validations, 'dataValidation') else 0
        print(f"Validation sayısı: {validation_count}")
    
    print("✅ Test başarılı!")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    traceback.print_exc()
    sys.exit(1)






