#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter

excel_path = r"C:\Users\Mert\Desktop\set\B10_OD_Dosya_Incelemesi.xlsx"

print("📂 Template Excel detaylı analiz...")

try:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb['B10_FileReview']
    
    print(f"\n📋 B10_FileReview Sayfası Detaylı Analiz:")
    print(f"   Toplam satır: {ws.max_row}")
    print(f"   Toplam sütun: {ws.max_column}")
    
    # Header satırını bul
    print(f"\n🔍 Header satırı analizi (Satır 1-10):")
    for row in range(1, 11):
        row_values = []
        for col in range(1, 11):
            cell = ws.cell(row, col)
            val = cell.value
            if val:
                row_values.append(f"{get_column_letter(col)}:{val}")
        if row_values:
            print(f"   Satır {row}: {', '.join(row_values)}")
    
    # İlk veri satırlarını analiz et
    print(f"\n📊 İlk 10 veri satırı (Satır 7-16):")
    for row in range(7, 17):
        row_data = []
        for col in range(1, 11):
            cell = ws.cell(row, col)
            val = cell.value
            if val or (col <= 3):  # İlk 3 sütunu her zaman göster
                row_data.append(f"{get_column_letter(col)}:{val if val else '(boş)'}")
        if any(v for v in row_data if '(boş)' not in v) or row <= 10:
            print(f"   Satır {row}: {', '.join(row_data)}")
    
    # B10 ve B10.01 referanslarını bul
    print(f"\n🔎 B10 ve B10.01 referanslarının konumları:")
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row, col)
            if cell.value and ('B10' in str(cell.value)):
                print(f"   Satır {row}, Sütun {get_column_letter(col)}: {cell.value}")
    
    # Soru numaralarını bul
    print(f"\n🔢 Soru numaralarının konumları:")
    for row in range(1, min(ws.max_row + 1, 50)):
        for col in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, (int, float)) and 1 <= cell.value <= 20:
                # Check if it's a question number (not just any number)
                next_cell = ws.cell(row, col + 1)
                if next_cell.value and isinstance(next_cell.value, str) and len(str(next_cell.value)) > 10:
                    print(f"   Satır {row}, Sütun {get_column_letter(col)}: Soru {int(cell.value)}")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    import traceback
    traceback.print_exc()


