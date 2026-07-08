import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys

excel_path = r"C:\Users\Mert\Desktop\set\B10_OD_Dosya_Incelemesi.xlsx"

try:
    print("📂 Excel dosyası analiz ediliyor...")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    print(f"✅ Workbook yüklendi. Sayfa sayısı: {len(wb.sheetnames)}")
    
    html_content = """<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Denetim Rehberi</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        body {
            background-color: #f5f7fa;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }
        .header-section {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 2rem 0;
            margin-bottom: 2rem;
        }
        .sheet-card {
            background: white;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            margin-bottom: 2rem;
            overflow: hidden;
        }
        .sheet-header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 1rem 1.5rem;
            font-weight: 600;
            font-size: 1.2rem;
        }
        .table-container {
            padding: 1.5rem;
            overflow-x: auto;
        }
        table {
            width: 100%;
            border-collapse: collapse;
        }
        th {
            background-color: #f8f9fa;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            border-bottom: 2px solid #dee2e6;
        }
        td {
            padding: 10px 12px;
            border-bottom: 1px solid #e9ecef;
        }
        tr:hover {
            background-color: #f8f9fa;
        }
        .validation-badge {
            display: inline-block;
            background: #28a745;
            color: white;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.85rem;
            margin-left: 8px;
        }
        .validation-cell {
            background-color: #e7f3ff !important;
            position: relative;
        }
        .validation-cell:hover {
            background-color: #d0e7ff !important;
        }
        .dropdown-icon {
            color: #28a745;
            margin-left: 5px;
        }
        .stats-card {
            background: white;
            border-radius: 10px;
            padding: 1.5rem;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            margin-bottom: 2rem;
        }
        .stat-item {
            text-align: center;
            padding: 1rem;
        }
        .stat-number {
            font-size: 2rem;
            font-weight: bold;
            color: #667eea;
        }
        .stat-label {
            color: #6c757d;
            margin-top: 0.5rem;
        }
        .tooltip {
            position: relative;
            display: inline-block;
        }
        .tooltip .tooltiptext {
            visibility: hidden;
            width: 300px;
            background-color: #333;
            color: #fff;
            text-align: left;
            border-radius: 6px;
            padding: 8px 12px;
            position: absolute;
            z-index: 1000;
            bottom: 125%;
            left: 50%;
            margin-left: -150px;
            opacity: 0;
            transition: opacity 0.3s;
            font-size: 0.85rem;
        }
        .tooltip:hover .tooltiptext {
            visibility: visible;
            opacity: 1;
        }
    </style>
</head>
<body>
    <div class="header-section">
        <div class="container">
            <h1><i class="fas fa-clipboard-check"></i> Denetim Rehberi</h1>
            <p class="mb-0">B10_OD_Dosya_Incelemesi.xlsx Analiz Raporu</p>
        </div>
    </div>

    <div class="container">
        <div class="row">
            <div class="col-md-12">
                <div class="stats-card">
                    <div class="row" id="stats-row">
                        <!-- Stats will be inserted here -->
                    </div>
                </div>
            </div>
        </div>
"""
    
    # Process each sheet
    total_sheets = len(wb.sheetnames)
    total_validations = 0
    total_rows = 0
    
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"📊 Sayfa {sheet_idx}/{total_sheets}: {sheet_name}")
        
        # Read data with pandas
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            df = df.fillna('')
            total_rows += len(df)
        except Exception as e:
            print(f"⚠️ Pandas okuma hatası: {e}")
            df = pd.DataFrame()
        
        # Find validations
        validations = {}
        
        try:
            for dv in ws.data_validations.dataValidation:
                formula1 = dv.formula1 if hasattr(dv, 'formula1') else ''
                dv_type = dv.type if hasattr(dv, 'type') else 'unknown'
                
                if dv_type == 'list' and formula1:
                    # Get list values
                    list_values = []
                    formula = formula1.replace('$', '').replace('"', '')
                    
                    if '!' in formula or ':' in formula:
                        try:
                            if '!' in formula:
                                ref_sheet_name, ref_range = formula.split('!')
                                ref_ws = wb[ref_sheet_name]
                            else:
                                ref_ws = ws
                                ref_range = formula
                            
                            if ':' in ref_range:
                                start, end = ref_range.split(':')
                                start_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, start)))
                                start_row = int(''.join(filter(str.isdigit, start)))
                                end_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, end)))
                                end_row = int(''.join(filter(str.isdigit, end)))
                                
                                for r in range(start_row, end_row + 1):
                                    for c in range(start_col, end_col + 1):
                                        val = ref_ws.cell(row=r, column=c).value
                                        if val is not None and str(val).strip():
                                            list_values.append(str(val).strip())
                            else:
                                val = ref_ws[ref_range].value
                                if val:
                                    list_values.append(str(val))
                        except Exception as e:
                            list_values = [formula]
                    else:
                        if ';' in formula:
                            list_values = [v.strip() for v in formula.split(';') if v.strip()]
                        elif ',' in formula:
                            list_values = [v.strip() for v in formula.split(',') if v.strip()]
                        else:
                            list_values = [formula] if formula else []
                    
                    # Get cells in range
                    for cell_range in dv.ranges:
                        range_str = str(cell_range)
                        
                        if ':' in range_str:
                            start, end = range_str.split(':')
                            start_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, start)))
                            start_row = int(''.join(filter(str.isdigit, start)))
                            end_col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, end)))
                            end_row = int(''.join(filter(str.isdigit, end)))
                            
                            for r in range(start_row, end_row + 1):
                                for c in range(start_col, end_col + 1):
                                    cell_ref = f"{get_column_letter(c)}{r}"
                                    validations[cell_ref] = {
                                        'values': list_values,
                                        'type': dv_type
                                    }
                        else:
                            validations[range_str] = {
                                'values': list_values,
                                'type': dv_type
                            }
        except Exception as e:
            print(f"⚠️ Validation hatası: {e}")
        
        total_validations += len(validations)
        print(f"   ✅ {len(validations)} çoktan seçmeli hücre bulundu")
        
        # Generate HTML for this sheet
        html_content += f"""
        <div class="sheet-card">
            <div class="sheet-header">
                <i class="fas fa-table"></i> {sheet_name}
                <span class="validation-badge">{len(validations)} Çoktan Seçmeli Hücre</span>
            </div>
            <div class="table-container">
                <table class="table">
                    <thead>
                        <tr>
        """
        
        # Add column headers
        if len(df) > 0 and len(df.columns) > 0:
            for col_idx in range(len(df.columns)):
                html_content += f'<th>{get_column_letter(col_idx + 1)}</th>'
        else:
            html_content += '<th>A</th>'
        
        html_content += """
                        </tr>
                    </thead>
                    <tbody>
        """
        
        # Add rows (limit to 200 rows for performance)
        max_rows = min(200, len(df)) if len(df) > 0 else 0
        for row_idx in range(max_rows):
            html_content += '<tr>'
            for col_idx in range(len(df.columns) if len(df) > 0 else 1):
                cell_ref = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                cell_value = str(df.iloc[row_idx, col_idx]) if row_idx < len(df) and col_idx < len(df.columns) else ''
                
                is_validation = cell_ref in validations
                validation_class = 'validation-cell' if is_validation else ''
                
                html_content += f'<td class="{validation_class}">'
                
                if is_validation:
                    validation_data = validations[cell_ref]
                    values_str = '<br>'.join(validation_data['values'][:10])
                    if len(validation_data['values']) > 10:
                        values_str += f'<br>... (+{len(validation_data["values"]) - 10} daha)'
                    
                    html_content += f'''
                    <span class="tooltip">
                        {cell_value}
                        <i class="fas fa-chevron-circle-down dropdown-icon"></i>
                        <span class="tooltiptext">
                            <strong>Çoktan Seçmeli Seçenekler:</strong><br>
                            {values_str}
                        </span>
                    </span>
                    '''
                else:
                    html_content += cell_value
                
                html_content += '</td>'
            html_content += '</tr>'
        
        html_content += """
                    </tbody>
                </table>
            </div>
        </div>
        """
    
    # Add stats
    html_content = html_content.replace(
        '<!-- Stats will be inserted here -->',
        f'''
        <div class="col-md-4">
            <div class="stat-item">
                <div class="stat-number">{total_sheets}</div>
                <div class="stat-label">Sayfa</div>
            </div>
        </div>
        <div class="col-md-4">
            <div class="stat-item">
                <div class="stat-number">{total_validations}</div>
                <div class="stat-label">Çoktan Seçmeli Hücre</div>
            </div>
        </div>
        <div class="col-md-4">
            <div class="stat-item">
                <div class="stat-number">{total_rows}</div>
                <div class="stat-label">Toplam Satır</div>
            </div>
        </div>
        '''
    )
    
    html_content += """
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""
    
    # Save HTML
    output_path = 'html/denetim-rehberi.html'
    os.makedirs('html', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"\n✅ HTML sayfası oluşturuldu: {output_path}")
    print(f"📊 İstatistikler:")
    print(f"   - Sayfa sayısı: {total_sheets}")
    print(f"   - Çoktan seçmeli hücre: {total_validations}")
    print(f"   - Toplam satır: {total_rows}")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
