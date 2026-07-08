import openpyxl
import pandas as pd
import os

excel_path = r"C:\Users\Mert\Desktop\set\B10_OD_Dosya_Incelemesi.xlsx"

print("Starting...")
print(f"Excel path exists: {os.path.exists(excel_path)}")

try:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    print(f"✅ Workbook loaded. Sheets: {wb.sheetnames}")
    
    # Test: Create simple HTML
    html = f"""<!DOCTYPE html>
<html>
<head><title>Test</title></head>
<body>
<h1>Test - {len(wb.sheetnames)} sheets found</h1>
</body>
</html>"""
    
    output_path = 'html/denetim-rehberi.html'
    os.makedirs('html', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"✅ HTML created at: {output_path}")
    print(f"File exists: {os.path.exists(output_path)}")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()






