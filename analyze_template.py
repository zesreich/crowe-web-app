#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import json
import os

excel_path = r"C:\Users\Mert\Desktop\set\B10_OD_Dosya_Incelemesi.xlsx"

print("📂 Template Excel dosyası analiz ediliyor...")

try:
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    print(f"✅ Workbook yüklendi. Sayfa sayısı: {len(wb.sheetnames)}")
    print(f"📄 Sayfalar: {', '.join(wb.sheetnames)}")
    
    template_info = {
        'sheets': [],
        'structure': {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📋 Sayfa: {sheet_name}")
        print(f"   Satır sayısı: {ws.max_row}")
        print(f"   Sütun sayısı: {ws.max_column}")
        
        # Analyze structure
        sheet_info = {
            'name': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'headers': [],
            'data_validations': [],
            'merged_cells': list(ws.merged_cells.ranges),
            'sample_data': []
        }
        
        # Get headers (first row)
        if ws.max_row > 0:
            headers = []
            for col in range(1, min(ws.max_column + 1, 10)):
                cell = ws.cell(row=1, column=col)
                headers.append({
                    'column': get_column_letter(col),
                    'value': cell.value,
                    'style': {
                        'font': str(cell.font),
                        'fill': str(cell.fill),
                        'alignment': str(cell.alignment)
                    }
                })
            sheet_info['headers'] = headers
        
        # Get sample data (first 5 rows)
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append({
                    'column': get_column_letter(col_idx),
                    'value': cell.value,
                    'data_type': type(cell.value).__name__
                })
            sheet_info['sample_data'].append(row_data)
        
        # Get data validations
        for validation in ws.data_validations.dataValidation:
            range_str = ''
            if validation.ranges:
                try:
                    range_str = str(list(validation.ranges)[0])
                except:
                    range_str = str(validation.ranges)
            sheet_info['data_validations'].append({
                'range': range_str,
                'type': validation.type,
                'formula1': validation.formula1 if hasattr(validation, 'formula1') else None
            })
        
        template_info['sheets'].append(sheet_info)
        
        # Print sample
        print(f"   İlk satır (header): {[h['value'] for h in headers[:5]]}")
    
    # Save template info to JSON
    template_json_path = 'html/template_structure.json'
    with open(template_json_path, 'w', encoding='utf-8') as f:
        json.dump(template_info, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Template analizi tamamlandı!")
    print(f"📄 Yapı bilgisi kaydedildi: {template_json_path}")
    
except Exception as e:
    print(f"❌ Hata: {e}")
    import traceback
    traceback.print_exc()

